from openpyxl import Workbook, load_workbook
import os
import requests
import re
import pandas as pd
import fitz  # PyMuPDF library for PDF parsing
from datetime import datetime
import urllib3
from openpyxl.styles import Font

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def fetch_pdf_urls(month, year, doc_type):
  base_url = "https://www.srpc.kar.nic.in/website/2023/commercial/"
  pdf_urls = []
  month_locations = {}

  month_location = ["p", "f"]
  for ml in month_location:
    url = f"{base_url}{doc_type.lower()}{month[:3].lower()}{year[-2:]}{ml}.pdf"

    with requests.get(url, verify=False) as response:
      if response.status_code == 200:
        pdf_urls.append(url)
        month_locations[url] = ml
  return pdf_urls, month_locations


def extract_text_from_pdf(pdf_content):
  with fitz.open("pdf", pdf_content) as doc:
    text = ""
    for page in doc:
      text += page.get_text()
  return text


def find_table(text):
  table_pattern_4col = r'Entity\s+Total\s+Energy\s+Schedule\s+Total\s+Actual\s+data\s+Net\s+Deviation\s+for\s+the\s+purpose\s+of\s+REC'
  if re.search(table_pattern_4col, text, re.MULTILINE):
    return 4
  table_pattern_2col = r'Entity[\s\n]*Total Energy(?: Schedule)?'
  if re.search(table_pattern_2col, text, re.MULTILINE):
    return 2
  return 0


def extract_data(text, num_columns):
  data = []
  date_pattern = r'Actual Meter Reading Available Upto : (\d{4}-\d{2}-\d{2})'
  date_match = re.search(date_pattern, text)
  date_str = date_match.group(1) if date_match else "N/A"

  processed_entities = set()

  for entity in search_terms:
    if entity in processed_entities:
      continue
    if num_columns == 2:
      entity_pattern = f'(?m)^{entity}\n(.*?)\n'
    elif num_columns == 4:
      entity_pattern = f'(?m)^{entity}\n(.*?)\n(.*?)\n(.*?)\n'

    entity_match = re.search(entity_pattern, text, re.MULTILINE)
    if entity_match:
      values = [entity]
      value_lines = entity_match.groups()
      # if num_columns == 2, append empty string for the next two columns
      if num_columns == 2:
        value_lines += ("N/A", ) * 2  # Convert list to tuple here

      values.extend(value_lines[:num_columns -
                                1])  # Extracting data based on num_columns
      # if num_columns == 2 then extend with empty two "N/A" values"
      if num_columns == 2:
        values.extend(["N/A", "N/A"])
      data.append(values)
      processed_entities.add(entity)
      print("values", values)
      print("value_line", value_lines)
      # print("components", components)
      print("num_columns", num_columns)

  columns = [
      "Entity", "Total Energy Schedule", "Total Actual data",
      "Net Deviation for the purpose of REC"
  ]
  d = pd.DataFrame(data, columns=columns)
  print("d", d)
  return d, date_str


if __name__ == "__main__":
  doc_type = "REA"
  month = "jan"
  year = "2024"
  search_terms = ["SPRNG,NPKUNTA", "Fortum Solar,PAVAGADA",
                  "SPRNG,PUGULUR"]  # entity names
  solar_entities = {"SPRNG,NPKUNTA", "Fortum Solar,PAVAGADA"}
  non_solar_entities = {"SPRNG,PUGULUR"}

  date_ = ""

  # Fetch PDF URLs
  pdf_urls, month_locations = fetch_pdf_urls(month, year, doc_type)

  solar_data, non_solar_data = [], []

  for pdf_url, month_location in zip(pdf_urls, month_locations.values()):
    with requests.get(pdf_url, verify=False) as response:
      text = extract_text_from_pdf(response.content)
      num_columns = find_table(text)
      if num_columns:
        data, date_str = extract_data(text, num_columns)
        date_ = date_str
        if data is not None:
          for row in data.values:
            entity = row[0]
            if entity in solar_entities:
              solar_data.append((*row, month_location, pdf_url))
            elif entity in non_solar_entities:
              non_solar_data.append((*row, month_location, pdf_url))
            else:
              print(
                  f"Entity '{entity}' not found in solar_entities or non_solar_entities sets."
              )

  # Define column names
  solar_columns = [
      "Entity", "Total Energy Schedule", "Total Actual data",
      "Net Deviation for the purpose of REC", "Month Location", "PDF URL"
  ]
  non_solar_columns = [
      "Entity", "Total Energy Schedule", "Total Actual data",
      "Net Deviation for the purpose of REC", "Month Location", "PDF URL"
  ]

  # Create DataFrames
  solar_df = pd.DataFrame(solar_data, columns=solar_columns)
  non_solar_df = pd.DataFrame(non_solar_data, columns=non_solar_columns)

  # Save DataFrames as CSV files
  # solar_df.to_csv("solar_data.csv", index=False)
  # non_solar_df.to_csv("non_solar_data.csv", index=False)

  print("Actual Meter Reading Available Upto :", date_)
  filename = f"Extracted Data_WRPC_SRPC_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
  sheet_name = "SRPC_REA"

  # Check file existence
  if not os.path.exists(filename):
      wb = Workbook()
      wb.save(filename)
  else:
      wb = load_workbook(filename)

  # Check if sheet exists
  if sheet_name not in wb.sheetnames:
      wb.create_sheet(title=sheet_name)
  ws2 = wb[sheet_name]

  ws2.cell(row=1, column=1, value="Actual Meter Reading Available Upto :")
  ws2.cell(row=1, column=2, value=date_)

  # Write headers for solar_df
  solar_headers = solar_df.columns
  for col_num, header in enumerate(solar_headers, start=1):
      cell = ws2.cell(row=2, column=col_num, value=header)
      cell.font = Font(bold=True)

  # Write solar_df data
  for row_num, (_, row_data) in enumerate(solar_df.iterrows(), start=3):
      for col_num, value in enumerate(row_data, start=1):
          ws2.cell(row=row_num, column=col_num, value=value)

  # Write headers for non_solar_df
  non_solar_headers = non_solar_df.columns
  for col_num, header in enumerate(non_solar_headers, start=1):
      cell = ws2.cell(row=len(solar_df) + 4, column=col_num, value=header)
      cell.font = Font(bold=True)

  # Write non_solar_df data
  for row_num, (_, row_data) in enumerate(non_solar_df.iterrows(), start=len(solar_df) + 5):
      for col_num, value in enumerate(row_data, start=1):
          ws2.cell(row=row_num, column=col_num, value=value)

  # Save the workbook
  wb.save(filename)
